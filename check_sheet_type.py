# -*- coding: utf-8 -*-
import openpyxl

def detect_sheet_type(sheet_name, headers):
    """ExcelService.DetectSheetType 시뮬레이션"""
    
    # 1. 시트명 기반 감지
    if '회생' in sheet_name or 'Sheet A-1' in sheet_name or 'A-1' in sheet_name:
        return 'BorrowerRestructuring'
    
    if '차주일반' in sheet_name or 'Sheet A' in sheet_name:
        return 'BorrowerGeneral'
    
    if '채권' in sheet_name or 'Sheet B' in sheet_name:
        return 'Loan'
    
    if '등기부등본' in sheet_name or 'Sheet C-2' in sheet_name or 'C-2' in sheet_name:
        return 'RegistryDetail'
    
    if '담보설정' in sheet_name or 'Sheet C-3' in sheet_name or 'C-3' in sheet_name:
        return 'CollateralSetting'
    
    if '물건정보' in sheet_name or 'Sheet C-1' in sheet_name or 'C-1' in sheet_name:
        return 'Property'
    
    if ('담보' in sheet_name or '물건' in sheet_name) and 'C-2' not in sheet_name and 'C-3' not in sheet_name:
        return 'Property'
    
    if '보증' in sheet_name or 'Sheet D' in sheet_name:
        return 'Guarantee'
    
    # 2. 헤더 패턴 fallback
    header_str = ' '.join(headers)
    if 'Property' in header_str or '담보소재지' in header_str or '경매사건번호' in header_str:
        return 'Property'
    
    return 'Unknown'

# Excel 파일
wb = openpyxl.load_workbook(r'C:\Users\pwm89\dev\nplogic\manager\client\IBK 2025-3 Program Data Disk I_Pool B.xlsx', read_only=True, data_only=True)

print('=== Sheet Type Detection ===')
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # 헤더 읽기 (Row 10 가정)
    headers = []
    for col in range(1, min(ws.max_column + 1, 30)):
        val = ws.cell(10, col).value
        if val:
            headers.append(str(val).replace('\n', ' '))
    
    sheet_type = detect_sheet_type(sheet_name, headers)
    print(f'{sheet_name}: {sheet_type}')
    
    # Property 시트만 상세 정보
    if sheet_type == 'Property':
        print(f'  -> Data rows: {ws.max_row - 10}')
