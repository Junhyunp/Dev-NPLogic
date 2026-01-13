# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\pwm89\dev\nplogic\manager\client\IBK 2025-3 Program Data Disk I_Pool B.xlsx', read_only=True, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n===== {sheet_name} =====')
    print(f'rows: {ws.max_row}, cols: {ws.max_column}')
    
    # 헤더 행 찾기
    header_row = None
    for row in range(1, min(15, ws.max_row + 1)):
        cell1 = str(ws.cell(row, 1).value or '')
        cell2 = str(ws.cell(row, 2).value or '')
        if 'Field' in cell1 or 'Field' in cell2:
            header_row = row + 1
            print(f'Found Field pattern at row {row}, header at {header_row}')
            break
    
    if header_row is None:
        for row in range(1, min(15, ws.max_row + 1)):
            cell2 = str(ws.cell(row, 2).value or '')
            if '일련번호' in cell2:
                header_row = row
                print(f'Found 일련번호 at row {row}')
                break
    
    if header_row:
        # 헤더 출력
        print(f'Header columns (first 15):')
        for col in range(1, min(ws.max_column + 1, 16)):
            val = ws.cell(header_row, col).value
            if val:
                val_str = str(val).replace('\n', ' ').replace('\r', '')
                print(f'  Col {col}: {val_str}')
        
        # 첫 데이터 행 출력
        data_row = header_row + 1
        if data_row <= ws.max_row:
            print(f'First data row (row {data_row}):')
            for col in range(1, min(ws.max_column + 1, 10)):
                val = ws.cell(data_row, col).value
                if val:
                    print(f'  Col {col}: {str(val)[:30]}')
        
        # 데이터 행 수
        print(f'Data rows: {ws.max_row - header_row}')
    else:
        print('Header row not found')
