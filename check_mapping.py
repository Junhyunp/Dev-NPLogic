# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\pwm89\dev\nplogic\manager\client\IBK 2025-3 Program Data Disk I_Pool B.xlsx', read_only=True, data_only=True)

ws = wb['Sheet C-1']
print('===== Sheet C-1 상세 분석 =====')
print(f'총 행: {ws.max_row}, 총 열: {ws.max_column}')

# Row 10이 헤더
header_row = 10
print(f'\n=== 모든 헤더 컬럼 (Row {header_row}) ===')
headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(header_row, col).value
    if val:
        val_str = str(val).replace('\n', ' ').replace('\r', '').strip()
        headers[col] = val_str
        print(f'Col {col}: "{val_str}"')

# 매핑 규칙과 비교
print('\n=== 매핑 확인 ===')
mapping_targets = {
    '차주일련번호': 'borrower_number',
    '차주명': 'borrower_name', 
    '물건번호': 'collateral_number',
    'Property 일련번호': 'property_serial',
    '경매사건번호': 'property_number',
    'Property Type': 'property_type',
}

for col_num, col_name in headers.items():
    for target, db_col in mapping_targets.items():
        if target in col_name:
            print(f'Col {col_num} "{col_name}" → {db_col}')

# 첫 5개 데이터 행
print('\n=== 첫 5개 데이터 행 ===')
key_cols = [5, 6, 7]  # 차주일련번호, 차주명, 물건번호
for row in range(11, 16):
    data = []
    for col in key_cols:
        val = ws.cell(row, col).value
        data.append(f'Col{col}={val}')
    print(f'Row {row}: {data}')
