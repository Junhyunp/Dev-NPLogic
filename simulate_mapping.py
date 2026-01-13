# -*- coding: utf-8 -*-
import openpyxl

# 매핑 규칙 (SheetMappingConfig.GetPropertyMappings 복사)
mapping_rules = [
    ('일련번호', 'serial_number'),
    ('Pool 구분', 'pool_type'),
    ('차주구분', 'borrower_category'),
    ('차주일련번호', 'borrower_number'),
    ('차주명', 'borrower_name'),
    ('물건번호', 'collateral_number'),
    ('Property 일련번호', 'property_serial'),
    ('Property일련번호', 'property_serial'),
    ('경매사건번호', 'property_number'),
    ('경매사건번호(IBK)', 'property_number'),
    ('경매사건번호 (IBK)', 'property_number'),
    ('Property Type', 'property_type'),
    ('Property-대지면적', 'land_area'),
    ('Property-건물면적', 'building_area'),
    ('담보소재지1', 'address_province'),
    ('담보소재지 1', 'address_province'),
    ('담보소재지2', 'address_city'),
    ('담보소재지 2', 'address_city'),
    ('담보소재지3', 'address_district'),
    ('담보소재지 3', 'address_district'),
    ('담보소재지4', 'address_detail'),
    ('담보소재지 4', 'address_detail'),
]

def normalize(name):
    """NormalizeColumnName 함수 시뮬레이션"""
    if not name:
        return ''
    return name.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ').replace('  ', ' ').strip()

def find_mapping(rules, excel_col_name):
    """FindMappingRule 함수 시뮬레이션"""
    if not excel_col_name:
        return None
    
    normalized = normalize(excel_col_name)
    
    # 1. 정확한 매칭
    for rule_name, db_col in rules:
        if normalize(rule_name).lower() == normalized.lower():
            return db_col
    
    # 2. 포함 매칭 (Excel 컬럼이 규칙 포함)
    for rule_name, db_col in rules:
        if normalize(rule_name).lower() in normalized.lower():
            return db_col
    
    # 3. 역방향 포함 매칭
    for rule_name, db_col in rules:
        if normalized.lower() in normalize(rule_name).lower():
            return db_col
    
    return None

# Excel 파일 읽기
wb = openpyxl.load_workbook(r'C:\Users\pwm89\dev\nplogic\manager\client\IBK 2025-3 Program Data Disk I_Pool B.xlsx', read_only=True, data_only=True)
ws = wb['Sheet C-1']

print('=== Sheet C-1 Mapping Simulation ===')
print()

# 헤더 Row 10
header_row = 10
headers = {}
for col in range(1, ws.max_column + 1):
    val = ws.cell(header_row, col).value
    if val:
        headers[col] = str(val)

# 매핑 테스트
print('Column Mapping Results:')
mapped_cols = {}
for col, header in headers.items():
    db_col = find_mapping(mapping_rules, header)
    if db_col:
        mapped_cols[col] = (header.replace('\n', '\\n'), db_col)
        print(f'  Col {col}: "{header.replace(chr(10), "\\n")}" -> {db_col}')

print()
print(f'Total mapped columns: {len(mapped_cols)}')

# 중요 컬럼 확인
print()
print('=== Key Columns Check ===')
important = ['borrower_number', 'borrower_name', 'collateral_number', 'property_number', 'property_type']
for db_col in important:
    found = [col for col, (h, d) in mapped_cols.items() if d == db_col]
    if found:
        print(f'{db_col}: Col {found}')
    else:
        print(f'{db_col}: NOT FOUND!')

# 첫 데이터 행으로 실제 값 확인
print()
print('=== First Data Row Values ===')
data_row = 11
for col, (header, db_col) in mapped_cols.items():
    if db_col in important:
        val = ws.cell(data_row, col).value
        print(f'{db_col} (Col {col}): {val}')
