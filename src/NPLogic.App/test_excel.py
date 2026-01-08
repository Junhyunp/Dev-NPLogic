# -*- coding: utf-8 -*-
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'C:\Users\pwm89\dev\nplogic\manager\client\IBK 2025-3 Program Data Disk I_Pool B.xlsx'
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

print('=== Sheet C-1 Row 10 (Header) - All columns ===')
ws = wb['Sheet C-1']
for idx, cell in enumerate(ws[10], 1):
    val = str(cell.value).replace('\n', ' ').replace('\r', '') if cell.value else ''
    if val:
        print(f'  Col {idx}: "{val}"')

print()
print('=== Sheet C-1 Row 11 (First Data) ===')
for idx, cell in enumerate(ws[11], 1):
    val = str(cell.value)[:50] if cell.value else ''
    if val:
        print(f'  Col {idx}: {val}')

print()
print('=== Sheet A Row 10 (Header) - All columns ===')
ws = wb['Sheet A']
for idx, cell in enumerate(ws[10], 1):
    val = str(cell.value).replace('\n', ' ').replace('\r', '') if cell.value else ''
    if val:
        print(f'  Col {idx}: "{val}"')

print()
print('=== Sheet A Row 11 (First Data) ===')
for idx, cell in enumerate(ws[11], 1):
    val = str(cell.value)[:50] if cell.value else ''
    if val:
        print(f'  Col {idx}: {val}')
