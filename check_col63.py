# -*- coding: utf-8 -*-
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\pwm89\dev\nplogic\manager\client\IBK 2025-3 Program Data Disk I_Pool B.xlsx', read_only=True, data_only=True)
ws = wb['Sheet C-1']

print('=== Col 63 data check ===')
header = ws.cell(10, 63).value
print(f'Header Col 63: {header}')

# Count data
has_data = 0
no_data = 0
for row in range(11, 125):
    val = ws.cell(row, 63).value
    if val:
        has_data += 1
    else:
        no_data += 1

print(f'Has data: {has_data}, No data: {no_data}')

# Sample
print('\nFirst 15 rows - Col 7 vs Col 63:')
for row in range(11, 26):
    col7 = ws.cell(row, 7).value   # 물건번호
    col63 = ws.cell(row, 63).value  # 경매사건번호
    print(f'Row {row}: col7={col7}, col63={col63}')
