# -*- coding: utf-8 -*-
"""
테스트 파일 생성 스크립트
- Excel 업로드 테스트용 샘플 물건 데이터
- PDF 업로드 테스트용 단순 문서
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# 현재 스크립트 위치
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def create_excel_file():
    """10개 샘플 물건 데이터가 포함된 Excel 파일 생성"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "물건목록"
    
    # 헤더 정의
    headers = [
        "ProjectId",
        "PropertyNumber", 
        "PropertyType",
        "AddressFull",
        "AddressRoad",
        "AddressJibun",
        "LandArea",
        "BuildingArea",
        "AppraisalValue",
        "MinimumBid"
    ]
    
    # 샘플 데이터 (10개)
    sample_data = [
        # ProjectId, PropertyNumber, PropertyType, AddressFull, AddressRoad, AddressJibun, LandArea, BuildingArea, AppraisalValue, MinimumBid
        ["PROJ-2024-001", "2024-001", "아파트", "서울특별시 강남구 테헤란로 123 OO아파트 101동 1001호", "서울특별시 강남구 테헤란로 123", "서울특별시 강남구 역삼동 123-45", 85.5, 112.3, 1500000000, 1200000000],
        ["PROJ-2024-001", "2024-002", "다가구", "서울특별시 서초구 방배동 456-78", "서울특별시 서초구 방배로 45", "서울특별시 서초구 방배동 456-78", 150.2, 320.5, 2800000000, 2240000000],
        ["PROJ-2024-001", "2024-003", "상가", "서울특별시 마포구 홍대입구역 근처 상가", "서울특별시 마포구 양화로 100", "서울특별시 마포구 서교동 100-1", 45.0, 55.8, 850000000, 680000000],
        ["PROJ-2024-001", "2024-004", "오피스텔", "경기도 성남시 분당구 정자동 XX타워 501호", "경기도 성남시 분당구 불정로 101", "경기도 성남시 분당구 정자동 5-1", 28.5, 32.1, 450000000, 360000000],
        ["PROJ-2024-001", "2024-005", "토지", "경기도 용인시 수지구 동천동 산 123", "경기도 용인시 수지구 동천로 200", "경기도 용인시 수지구 동천동 산 123", 500.0, 0, 3200000000, 2560000000],
        ["PROJ-2024-001", "2024-006", "아파트", "부산광역시 해운대구 우동 센텀시티 XX아파트 2502호", "부산광역시 해운대구 센텀중앙로 55", "부산광역시 해운대구 우동 1500", 120.8, 145.2, 980000000, 784000000],
        ["PROJ-2024-001", "2024-007", "공장", "경기도 화성시 동탄산업단지 내", "경기도 화성시 동탄산단1길 50", "경기도 화성시 석우동 200", 2000.5, 1500.0, 5500000000, 4400000000],
        ["PROJ-2024-001", "2024-008", "다세대", "인천광역시 남동구 구월동 XX빌라 301호", "인천광역시 남동구 구월로 88", "인천광역시 남동구 구월동 88-1", 65.3, 78.9, 320000000, 256000000],
        ["PROJ-2024-001", "2024-009", "근린시설", "대전광역시 서구 둔산동 타임빌딩 1층", "대전광역시 서구 둔산로 123", "대전광역시 서구 둔산동 100", 180.0, 200.5, 1200000000, 960000000],
        ["PROJ-2024-001", "2024-010", "아파트", "대구광역시 수성구 범어동 XX아파트 1801호", "대구광역시 수성구 동대구로 300", "대구광역시 수성구 범어동 50", 95.2, 118.6, 720000000, 576000000],
    ]
    
    # 스타일 정의
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 헤더 작성
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 데이터 작성
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # 숫자 컬럼 우측 정렬
            if col_idx >= 7:  # LandArea 이후
                cell.alignment = Alignment(horizontal='right')
                if col_idx >= 9:  # 금액 컬럼 서식
                    cell.number_format = '#,##0'
    
    # 컬럼 너비 조정
    column_widths = [18, 15, 12, 50, 35, 35, 12, 12, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # 파일 저장
    file_path = os.path.join(SCRIPT_DIR, "test_물건목록_샘플.xlsx")
    wb.save(file_path)
    print(f"Excel 파일 생성 완료: {file_path}")
    return file_path


def create_pdf_file():
    """단순 텍스트 PDF 파일 생성"""
    
    file_path = os.path.join(SCRIPT_DIR, "test_sample_document.pdf")
    
    # PDF 생성
    c = canvas.Canvas(file_path, pagesize=A4)
    width, height = A4
    
    # 기본 폰트 사용 (한글 미지원이므로 영어로 작성)
    c.setFont("Helvetica-Bold", 24)
    c.drawCentredString(width/2, height - 100, "Test Document")
    
    c.setFont("Helvetica", 14)
    c.drawCentredString(width/2, height - 150, "NPLogic PDF Upload Test")
    
    c.setFont("Helvetica", 12)
    y_position = height - 220
    
    test_content = [
        "This is a sample PDF document for testing file upload functionality.",
        "",
        "Purpose: Supabase Storage Upload Test",
        "Created: 2025-01-04",
        "File Type: PDF",
        "",
        "Test Information:",
        "- This document contains simple text content",
        "- Used for DataUploadView PDF upload testing",
        "- File will be stored in Supabase Storage 'documents' bucket",
        "",
        "Lorem ipsum dolor sit amet, consectetur adipiscing elit.",
        "Sed do eiusmod tempor incididunt ut labore et dolore magna aliqua.",
        "Ut enim ad minim veniam, quis nostrud exercitation ullamco laboris.",
    ]
    
    for line in test_content:
        c.drawString(72, y_position, line)
        y_position -= 20
    
    # 페이지 하단 정보
    c.setFont("Helvetica", 10)
    c.drawCentredString(width/2, 50, "NPLogic Test Document - Page 1 of 1")
    
    c.save()
    print(f"PDF 파일 생성 완료: {file_path}")
    return file_path


if __name__ == "__main__":
    print("=" * 50)
    print("테스트 파일 생성 시작")
    print("=" * 50)
    
    # Excel 파일 생성
    excel_path = create_excel_file()
    
    # PDF 파일 생성
    pdf_path = create_pdf_file()
    
    print("=" * 50)
    print("모든 테스트 파일 생성 완료!")
    print("=" * 50)

