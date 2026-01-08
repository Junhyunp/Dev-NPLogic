"""
IBK Multi-Sheet Upload 테스트용 Excel 파일 생성
4개 시트 구조: Sheet A, Sheet A-1, Sheet B, Sheet C-1
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime, timedelta
import random

def create_test_excel():
    wb = openpyxl.Workbook()
    
    # 스타일 정의
    header_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==================== Sheet A: 차주일반정보 ====================
    ws_a = wb.active
    ws_a.title = "차주일반정보"
    
    # 헤더 (IBK 형식 - 상단에 메타 정보)
    ws_a['B2'] = "IBK 2025-3 Program"
    ws_a['B3'] = "Pool B"
    ws_a['B4'] = "Sheet A"
    ws_a['B5'] = "차주일반정보"
    ws_a['B2'].font = Font(bold=True, italic=True, size=14)
    
    # 자산확정일 등 메타 정보
    ws_a['G2'] = "자산확정일"
    ws_a['H2'] = "2025-07-31"
    
    # 데이터 헤더 (10행)
    headers_a = ["일련번호", "Pool 구분", "차주구분", "차주일련번호", "차주명", "관련차주", 
                 "차주형태", "대출원금잔액", "가지급금", "미상환원금잔액", "미수이자", 
                 "차주별 근저당권설정액", "차주별 선순위 근저당설정액", "비고"]
    
    for col, header in enumerate(headers_a, 1):
        cell = ws_a.cell(row=10, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # 샘플 데이터
    borrowers = [
        (1, "B", "Regular", "R-001", "(주)테스트기업", "", "중소법인", 5000000000, 50000000, 4900000000, 100000000, 600000000, 500000000, ""),
        (2, "B", "Regular", "R-002", "홍길동", "", "개인", 1000000000, 10000000, 990000000, 20000000, 150000000, 100000000, ""),
        (3, "B", "Regular", "R-003", "(주)샘플회사", "", "중소법인", 3000000000, 30000000, 2950000000, 50000000, 400000000, 300000000, ""),
        (4, "B", "Special", "S-001", "(주)회생기업A", "", "중소법인", 2000000000, 20000000, 1980000000, 30000000, 250000000, 200000000, "회생진행중"),
        (5, "B", "Special", "S-002", "김회생", "", "개인", 500000000, 5000000, 495000000, 8000000, 80000000, 50000000, "회생진행중"),
    ]
    
    for row_idx, data in enumerate(borrowers, 11):
        for col_idx, value in enumerate(data, 1):
            cell = ws_a.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # ==================== Sheet A-1: 회생차주정보 ====================
    ws_a1 = wb.create_sheet("회생차주정보")
    
    ws_a1['B2'] = "IBK 2025-3 Program"
    ws_a1['B3'] = "Pool B"
    ws_a1['B4'] = "Sheet A-1"
    ws_a1['B5'] = "회생차주정보"
    ws_a1['B2'].font = Font(bold=True, italic=True, size=14)
    
    headers_a1 = ["일련번호", "Pool 구분", "차주구분", "차주일련번호", "차주명", "관련차주",
                  "인가/미인가", "세부진행단계", "관할법원", "회생사건번호", 
                  "회생신청일", "보전처분일", "개시결정일", "채권신고일", "인가/폐지결정일", "회생탈락권"]
    
    for col, header in enumerate(headers_a1, 1):
        cell = ws_a1.cell(row=10, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # 회생 차주 데이터
    restructuring = [
        (1, "B", "Special", "S-001", "(주)회생기업A", "", "미인가", "개시결정", "서울회생법원", "2025회합191", 
         datetime(2025, 5, 29), datetime(2025, 5, 30), datetime(2025, 7, 7), datetime(2025, 7, 7), None, "추후제공"),
        (2, "B", "Special", "S-002", "김회생", "", "미인가", "개시결정", "부산회생법원", "2024회합100096",
         datetime(2024, 7, 15), datetime(2024, 7, 17), datetime(2024, 7, 31), datetime(2024, 7, 31), None, "추후제공"),
    ]
    
    for row_idx, data in enumerate(restructuring, 11):
        for col_idx, value in enumerate(data, 1):
            cell = ws_a1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if isinstance(value, datetime):
                cell.number_format = 'YYYY-MM-DD'
    
    # ==================== Sheet B: 채권일반정보 ====================
    ws_b = wb.create_sheet("채권일반정보")
    
    ws_b['B2'] = "IBK 2025-3 Program"
    ws_b['B3'] = "Pool B"
    ws_b['B4'] = "Sheet B"
    ws_b['B5'] = "채권일반정보"
    ws_b['B2'].font = Font(bold=True, italic=True, size=14)
    
    headers_b = ["일련번호", "Pool 구분", "차주구분", "차주일련번호", "차주명", "대출일련번호",
                 "대출과목", "계좌번호", "이자율", "최초대출일", "대출만기일", "최종이수일",
                 "통화표시", "최초대출금액", "통화표시", "대출금잔액"]
    
    for col, header in enumerate(headers_b, 1):
        cell = ws_b.cell(row=10, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # 대출 데이터
    loans = [
        (1, "B", "Regular", "R-001", "(주)테스트기업", 1, "중소기업시설자금대출", "054**1884**034", 0.0872, 
         datetime(2018, 10, 5), datetime(2025, 9, 5), datetime(2024, 12, 4), "KRW", 268000000, "KRW", 180971638),
        (2, "B", "Regular", "R-001", "(주)테스트기업", 2, "중소기업자금대출", "054**1884**035", 0.0892,
         datetime(2018, 12, 21), datetime(2024, 12, 21), datetime(2024, 12, 22), "KRW", 500000000, "KRW", 500000000),
        (3, "B", "Regular", "R-002", "홍길동", 1, "개인사업자대출", "054**1884**041", 0.0931,
         datetime(2019, 3, 29), datetime(2025, 3, 29), datetime(2024, 11, 28), "KRW", 1350000000, "KRW", 1350000000),
        (4, "B", "Special", "S-001", "(주)회생기업A", 1, "중소기업시설자금대출", "054**1884**048", 0.0816,
         datetime(2019, 11, 18), datetime(2025, 10, 23), datetime(2024, 11, 22), "KRW", 2200000000, "KRW", 2200000000),
    ]
    
    for row_idx, data in enumerate(loans, 11):
        for col_idx, value in enumerate(data, 1):
            cell = ws_b.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if isinstance(value, datetime):
                cell.number_format = 'YYYY-MM-DD'
            elif isinstance(value, float) and value < 1:  # 이자율
                cell.number_format = '0.00%'
    
    # ==================== Sheet C-1: 담보재산정보_물건정보 ====================
    ws_c1 = wb.create_sheet("담보재산정보_물건정보")
    
    ws_c1['B2'] = "IBK 2025-3 Program"
    ws_c1['B3'] = "Pool B"
    ws_c1['B4'] = "Sheet C-1"
    ws_c1['B5'] = "담보재산정보_물건정보"
    ws_c1['B2'].font = Font(bold=True, italic=True, size=14)
    
    headers_c1 = ["일련번호", "Pool 구분", "차주구분", "차주일련번호", "차주명", "물건번호",
                  "Property 일련번호", "담보소재지1", "담보소재지2", "담보소재지3", 
                  "담보소재지4", "Property Type", "Property-대지면적", "Property-건물면적"]
    
    for col, header in enumerate(headers_c1, 1):
        cell = ws_c1.cell(row=10, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # 담보물건 데이터
    properties = [
        (1, "B", "Regular", "R-001", "(주)테스트기업", "R-001-1", "R-001-1", "경기도", "용인시 처인구", "포곡읍", 
         "56-3(토지), 1동, 2동, 3동, 4동 건물 4개동, 56-22, 56-23(토지 2필지)", "근린주택", 7934.00, 4091.03),
        (2, "B", "Regular", "R-001", "(주)테스트기업", "R-001-2", "R-001-2", "경기도", "화성시", "새솔동", 
         "테스트아파트 101동 1801호", "단독주택", 705.60, 567.83),
        (3, "B", "Regular", "R-002", "홍길동", "R-002-1", "R-002-1", "경기도", "김포시", "양촌읍 합은리", 
         "258-77, 258-107(토지 2필지, 양지A삼 건물)", "단독주택", 541.00, 185.85),
        (4, "B", "Regular", "R-003", "(주)샘플회사", "R-003-1", "R-003-1", "인천광역시", "서구", "청라동", 
         "15-20(토지 건물)", "근린주택", 365.60, 306.98),
        (5, "B", "Special", "S-001", "(주)회생기업A", "S-001-1", "S-001-1", "인천광역시", "서구", "청라동", 
         "164-1 청라솔라닉스타워 제10층 제29호 제290호1", "오피스텔(수익)", 48.25, 112.36),
    ]
    
    for row_idx, data in enumerate(properties, 11):
        for col_idx, value in enumerate(data, 1):
            cell = ws_c1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # 컬럼 너비 자동 조정
    for ws in [ws_a, ws_a1, ws_b, ws_c1]:
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 30)
    
    # 파일 저장
    output_path = "IBK_MultiSheet_Test.xlsx"
    wb.save(output_path)
    print(f"테스트 파일 생성 완료: {output_path}")
    print(f"\n포함된 시트:")
    print(f"  - 차주일반정보: 5명의 차주 (Regular 3, Special 2)")
    print(f"  - 회생차주정보: 2명의 회생 차주")
    print(f"  - 채권일반정보: 4건의 대출")
    print(f"  - 담보재산정보_물건정보: 5건의 담보물건")
    
    return output_path

if __name__ == "__main__":
    create_test_excel()


