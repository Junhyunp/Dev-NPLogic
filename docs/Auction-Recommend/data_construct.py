# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
import os
import re
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Tuple, Set
from tqdm import tqdm

import pandas as pd
import requests
import yaml
from bs4 import BeautifulSoup


# -----------------------------
# 사용처별 정규화/매핑 설정
# -----------------------------

# 키워드 → 정규화된 물건 종류(정확한 대표명)
# 긴 키워드부터 탐색해 겹침 방지(예: "아파트형공장" vs "공장")
_KEYWORD_TO_CANONICAL: List[Tuple[str, str]] = [
    ("아파트형공장", "아파트형공장"),
    ("다가구", "다가구(원룸등)"),
    ("원룸", "다가구(원룸등)"),
    ("다세대", "다세대(빌라)"),
    ("근린주택", "근린주택"),
    ("근린상가", "근린상가"),
    ("근린시설", "근린시설"),
    ("오피스텔", "오피스텔"),
    ("사무실", "사무실"),
    ("숙박시설", "숙박시설"),
    ("콘도", "숙박(콘도등)"),
    ("숙박", "숙박(콘도등)"),
    ("교육시설", "교육시설"),
    ("종교시설", "종교시설"),
    ("농가관련시설", "농가관련시설"),
    ("의료시설", "의료시설"),
    ("주유소(위험물)", "주유소(위험물)"),
    ("주유소", "주유소(위험물)"),
    ("위험물", "주유소(위험물)"),
    ("목욕탕", "목욕탕"),
    ("노유자시설", "노유자시설"),
    ("분뇨쓰레기처리", "분뇨쓰레기처리"),
    ("자동차관련시설", "자동차관련시설"),
    ("장례관련시설", "장례관련시설"),
    ("문화및집회시설", "문화및집회시설"),
    ("공장용지", "공장용지"),
    ("학교용지", "학교용지"),
    ("창고용지", "창고용지"),
    ("체육용지", "체육용지"),
    ("종교용지", "종교용지"),
    ("기타용지", "기타용지"),
    ("목장용지", "목장용지"),
    ("승용차", "승용차"),
    ("화물차", "화물차"),
    ("중장비", "중장비"),
    ("광업권", "광업권"),
    ("어업권", "어업권"),
    ("양어장", "양어장(축양,양식)"),
    ("축양", "양어장(축양,양식)"),
    ("양식", "양어장(축양,양식)"),
    ("염전", "염전"),
    ("주차장", "주차장"),
    ("대지", "대지"),
    ("농지", "농지"),
    ("임야", "임야"),
    ("잡종지", "잡종지"),
    ("과수원", "과수원"),
    ("도로", "도로"),
    ("묘지", "묘지"),
    ("구거", "구거"),
    ("하천", "하천"),
    ("유지", "유지"),
    ("제방", "제방"),
    ("창고", "창고"),
    ("공장", "공장"),
    ("아파트", "아파트"),
    ("주택", "주택"),
    ("상가", "근린상가"),  # 폭넓은 표현 방어
    ("기타", "기타"),
]


# 정규화된 물건 종류 → 그룹 번호(1~4)
# 주어진 이미지 기준으로 구성. 필요 시 수정 용이.
_CANONICAL_TO_GROUP: Dict[str, int] = {
    # 1. 아파트 그룹
    "아파트": 1,
    "오피스텔": 1,
    
    # 2. 연립다세대 그룹
    "다세대(빌라)" : 2,

    # 3. 공장/창고 그룹
    "공장": 3,
    "창고": 3,
    "농가관련시설": 3,
    "분뇨쓰레기처리": 3,
    # (필요 시 여기에 2그룹 항목 추가)

    # 4. 상가/아파트형공장 그룹
    "근린상가": 4,
    "사무실": 4,    
    "아파트형공장": 4,
    "숙박시설": 3,  
    "숙박(콘도등)": 4,
    "교육시설": 4,
    "종교시설": 4,

    # 5. 기본 그룹(주택, 근린시설, 토지 등 대부분)
    "주택": 5,
    "다가구(원룸등)": 5,
    "근린주택": 5,
    "근린시설": 5,
    "의료시설": 5,
    "주유소(위험물)": 5,
    "목욕탕": 5,
    "노유자시설": 5,
    "자동차관련시설": 5,
    "장례관련시설": 5,
    "문화및집회시설": 5,
    "대지": 5,
    "농지": 5,
    "임야": 5,
    "잡종지": 5,
    "과수원": 5,
    "도로": 5,
    "묘지": 5,
    "목장용지": 5,
    "공장용지": 5,
    "학교용지": 5,
    "창고용지": 5,
    "체육용지": 5,
    "종교용지": 5,
    "기타용지": 5,
    "구거": 5,
    "하천": 5,
    "유지": 5,
    "제방": 5,
    "주차장": 5,
    "승용차": 5,
    "버스": 5,
    "화물차": 5,
    "중장비": 5,
    "선박": 5,
    "광업권": 5,
    "어업권": 5,
    "염전": 5,
    "양어장(축양,양식)": 5,
    "기타": 5,
}


_BRACKET_PATTERN = re.compile(r"\[[^\]]*\]")
_SPACES_PATTERN = re.compile(r"\s+")


def _strip_brackets(text: str) -> str:
    cleaned = _BRACKET_PATTERN.sub("", text or "")
    cleaned = cleaned.replace("(", " (").replace(")", ") ")  # 토큰 경계 보조
    cleaned = _SPACES_PATTERN.sub(" ", cleaned)
    return cleaned.strip()


def normalize_usage(raw_usage: Optional[str]) -> Optional[str]:
    if not raw_usage:
        return None
    text = _strip_brackets(str(raw_usage))
    for keyword, canonical in _KEYWORD_TO_CANONICAL:
        if keyword in text:
            return canonical
    return "기타"


def usage_to_group(canonical_usage: Optional[str]) -> Optional[int]:
    if not canonical_usage:
        return None
    return _CANONICAL_TO_GROUP.get(canonical_usage, 4)


# 간단 문자열 유틸 (data_update.py와 동일한 로직)
def _to_str(x: object) -> str:
    try:
        s = str(x).strip()
        return "" if s.lower() == "nan" else s
    except Exception:
        return ""


# 간단 주소 파서(요청안): 특별시/광역시/도, 시/군/구, 읍/면/동/리
_ADDR_RE = re.compile(
    r"(?P<province>[가-힣]+?(특별시|광역시|도))?\s*"
    r"(?P<city>[가-힣]+?(시|군|구))?\s*"
    r"(?P<town>[가-힣0-9]+?(읍|면|동|리))?"
)

def parse_korean_address(address: Optional[str]) -> Optional[Dict[str, Optional[str]]]:
    if not address:
        return None
    s = str(address).strip()
    if not s:
        return None
    m = _ADDR_RE.search(s)
    return m.groupdict() if m else None


def _is_suspect_address(s: str) -> bool:
    """주소가 비정상/부정확해 보이는지 휴리스틱 판정 (data_update.py와 동일)."""
    ss = _to_str(s)
    if not ss:
        return True
    low = ss.lower()
    # 비주소 키워드(페이지 문구/지형/시설 등)
    bad_keywords = [
        "를 관할하는 시", "공시기준 공시가격",
        "주차장", "버스정류장", "포구", "해상", "부두",
    ]
    if any(k in ss for k in bad_keywords):
        return True
    # 정상적인 전체 주소는 보통 시/군/구가 포함된다.
    if ("시" in ss) or ("군" in ss) or ("구" in ss):
        return False
    # 시/군/구가 전혀 없고, 동/리/가 단위도 없는 경우(예: '현대아파트 108동 6층 603호')는 의심
    if not any(x in ss for x in ["동", "리", "가"]):
        return True
    # 길이가 비정상적으로 짧으면 의심
    if len(ss) <= 5:
        return True
    return True  # 시/군/구 없는 나머지 패턴은 보수적으로 의심 처리


def _is_row_suspect_address(row: "pd.Series") -> bool:
    """address 또는 address_road 중 하나라도 이상하면 True (data_update.py와 동일)."""
    a = _to_str(row.get("address"))
    r = _to_str(row.get("address_road"))
    return _is_suspect_address(a) or _is_suspect_address(r)


# -----------------------------
# VWorld 주소 → 좌표 (EPSG:4326)
# -----------------------------
_VWORLD_URL = "https://api.vworld.kr/req/address"
_VWORLD_DEFAULT_KEY = os.environ.get("VWORLD_KEY", "FA8A2F9F-12A3-35E7-9397-AD6EB13529C8")
_GEOCODE_CACHE: Dict[str, Tuple[Optional[float], Optional[float], Optional[str]]] = {}

def _parse_vworld_point(js: Dict[str, Any]) -> Tuple[Optional[float], Optional[float]]:
    try:
        res = js.get("response", {})
        if str(res.get("status")) != "OK":
            return None, None
        result = res.get("result")
        if isinstance(result, list) and result:
            point = result[0].get("point", {})
        else:
            point = (result or {}).get("point", {})
        x = point.get("x")
        y = point.get("y")
        return (float(x), float(y)) if x is not None and y is not None else (None, None)
    except Exception:
        return None, None

def geocode_address_vworld(address: Optional[str], prefer_type: str = "ROAD") -> Tuple[Optional[float], Optional[float], Optional[str]]:
    """문자열 주소를 VWorld로 좌표(EPSG:4326)로 변환.
    - prefer_type: "ROAD" 또는 "PARCEL" 우선 시도 후 실패 시 다른 타입 폴백
    - 반환: (lon, lat, "EPSG:4326") 또는 (None, None, None)
    캐시 사용으로 동일 주소 중복 호출 방지.
    """
    if not address:
        return None, None, None
    addr = str(address).strip()
    if not addr:
        return None, None, None
    if addr in _GEOCODE_CACHE:
        return _GEOCODE_CACHE[addr]

    key = _VWORLD_DEFAULT_KEY
    types = (prefer_type, "PARCEL" if prefer_type == "ROAD" else "ROAD")
    for ty in types:
        try:
            params = {
                "service": "address",
                "request": "getCoord",
                "crs": "epsg:4326",
                "format": "json",
                "type": ty,
                "address": addr,
                "key": key,
            }
            r = requests.get(_VWORLD_URL, params=params, timeout=6)
            if r.status_code != 200:
                continue
            x, y = _parse_vworld_point(r.json())
            if x is not None and y is not None:
                out = (x, y, "EPSG:4326")
                _GEOCODE_CACHE[addr] = out
                return out
        except Exception:
            continue
    out = (None, None, None)
    _GEOCODE_CACHE[addr] = out
    return out


def parse_date_to_iso(date_str: Optional[str]) -> Optional[str]:
    if not date_str:
        return None
    s = str(date_str).strip()
    if not s:
        return None
    s = s.replace("/", "-").replace(".", "-")
    # 허용 포맷: YYYY-MM-DD 또는 YYYY-MM-DD HH:MM:SS
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            continue
    # 추가 폴백: 길이 10 이상이면 앞 10자리만 시도
    try:
        return datetime.fromisoformat(s[:10]).strftime("%Y-%m-%d")
    except Exception:
        return s  # 원문 유지(엑셀 가독 목적)


def to_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(str(value).replace(",", "").strip())
    except Exception:
        try:
            return int(float(str(value).replace(",", "").strip()))
        except Exception:
            return None


def to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except Exception:
        return None


def iter_json_files(root_dir: str) -> Iterable[Tuple[str, str, str]]:
    """data/<court>/<area>/<n>.json 형태만 순회.

    반환: (abs_path, court, area)
    """
    root_dir = os.path.abspath(root_dir)
    for court in sorted(os.listdir(root_dir)):
        court_path = os.path.join(root_dir, court)
        if not os.path.isdir(court_path):
            continue
        for area in sorted(os.listdir(court_path)):
            area_path = os.path.join(court_path, area)
            if not os.path.isdir(area_path):
                continue
            for name in sorted(os.listdir(area_path)):
                if not name.lower().endswith(".json"):
                    continue
                yield os.path.join(area_path, name), court, area


# -----------------------------
# 진행상황(merge_progress.yaml) 지원
# -----------------------------
def _load_progress(progress_path: str) -> Dict[str, Any]:
    if not progress_path or not os.path.exists(progress_path):
        return {"completed": []}
    with open(progress_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    if "completed" not in data or not isinstance(data.get("completed"), list):
        data["completed"] = []
    return data


def _save_progress(progress_path: str, data: Dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(progress_path)) or ".", exist_ok=True)
    with open(progress_path, "w", encoding="utf-8") as f:
        yaml.safe_dump(data, f, allow_unicode=True, sort_keys=False)


def _is_completed(court: str, area: str, progress: Dict[str, Any]) -> bool:
    comp = progress.get("completed") or []
    for item in comp:
        if isinstance(item, str):
            if item == court:
                return True
        elif isinstance(item, dict):
            c = str(item.get("court") or "")
            a = str(item.get("area") or "")
            if c and a:
                if (c == court) and (a == area):
                    return True
            elif c and not a:
                if c == court:
                    return True
    return False


def _append_completed(progress: Dict[str, Any], pairs: Iterable[Tuple[str, str]]) -> Dict[str, Any]:
    comp = progress.get("completed") or []
    existing_pairs: Set[Tuple[str, str]] = set()
    existing_courts: Set[str] = set()
    for item in comp:
        if isinstance(item, str):
            existing_courts.add(item)
        elif isinstance(item, dict):
            c = str(item.get("court") or "")
            a = str(item.get("area") or "")
            if c and a:
                existing_pairs.add((c, a))
            elif c and not a:
                existing_courts.add(c)

    # 한 실행 내에서도 중복 방지
    seen_now: Set[Tuple[str, str]] = set()
    for (court, area) in pairs:
        key = (court, area)
        if key in seen_now:
            continue
        seen_now.add(key)
        if court in existing_courts:
            continue
        if (court, area) in existing_pairs:
            continue
        comp.append({"court": court, "area": area})
        existing_pairs.add((court, area))
    progress["completed"] = comp
    return progress


def build_row(doc: Dict[str, Any], court: str, area: str) -> Dict[str, Any]:
    # 키 보정: 사용승인일 철자 두 가지 모두 대응
    use_approval = (
        doc.get("use_approval_date")
        or doc.get("use_approaval_date")  # 일부 데이터 오타 대응
    )

    usage_canonical = normalize_usage(doc.get("usage"))
    group_no = usage_to_group(usage_canonical)

    # 주소 선택: new_address가 있으면 우선 사용
    chosen_address = doc.get("new_address") or doc.get("address")
    # 새 주소 파서 적용
    parsed = parse_korean_address(chosen_address) or {}
    big_region = parsed.get("province")
    mid_region = parsed.get("city")
    small_region = parsed.get("town")

    # 좌표(geocoding): VWorld 주소→좌표 API 사용 (EPSG:4326)
    lon, lat, crs = geocode_address_vworld(chosen_address)

    return {
        # 메타
        "auctionone_court": doc.get("court") or court,
        "auctionone_area": doc.get("area") or area,
        "region_big": big_region,
        "region_mid": mid_region,
        "region_small": small_region,

        # 핵심 식별/주소
        "case_no": doc.get("case_no"),
        "usage": usage_canonical,
        "object_group": group_no,  # 1~4
        "address": chosen_address,
        "address_road": None,
        "longitude": lon,
        "latitude": lat,
        "coord_crs": crs,

        # 일정
        "auction_date": parse_date_to_iso(doc.get("auction_date")),
        "big_round": to_int(doc.get("big_round")),

        # 금액
        "winning_price": to_int(doc.get("winning_price")),
        "second_big_price": to_int(doc.get("second_big_price")),

        # 감정/면적/기타 일자
        "valuation_base_date": parse_date_to_iso(doc.get("valuation_base_date")),
        "new_build_date": parse_date_to_iso(doc.get("new_build_date")),
        "use_approval_date": parse_date_to_iso(use_approval),

        "building_appraisal_price": to_int(doc.get("building_appraisal_price")),
        "land_appraisal_price": to_int(doc.get("land_appraisal_price")),
        "land_area": to_float(doc.get("land_area")),
        "building_area": to_float(doc.get("building_area")),

        # 링크
        "popup_url": doc.get("popup_url"),
    }


def construct_dataframe(input_root: str, use_progress: bool = False, progress_path: Optional[str] = None, update_progress: bool = False) -> Tuple[pd.DataFrame, List[Tuple[str, str]]]:
    rows: List[Dict[str, Any]] = []
    files = list(iter_json_files(input_root))
    processed_pairs_set: Set[Tuple[str, str]] = set()

    progress = {"completed": []}
    if use_progress:
        progress_path = progress_path or os.path.join(os.path.abspath(input_root), "merge_progress.yaml")
        progress = _load_progress(progress_path)

    def _to_scan(path: str, court: str, area: str) -> bool:
        if not use_progress:
            return True
        return not _is_completed(court, area, progress)

    for path, court, area in tqdm(files, desc="Scanning JSON", unit="file"):
        if not _to_scan(path, court, area):
            continue
        try:
            with open(path, "r", encoding="utf-8") as f:
                doc = json.load(f)
        except Exception:
            # 손상된 파일은 스킵
            continue
        row = build_row(doc, court, area)
        rows.append(row)
        # court/area는 JSON 파일 수와 무관하게 유일 쌍으로만 기록
        processed_pairs_set.add((court, area))

    if use_progress and update_progress and processed_pairs_set:
        try:
            progress_path2 = progress_path or os.path.join(os.path.abspath(input_root), "merge_progress.yaml")
            cur = _load_progress(progress_path2)
            # 이번 실행에서 처리된 유일 court/area만 기록
            cur = _append_completed(cur, list(processed_pairs_set))
            _save_progress(progress_path2, cur)
        except Exception as e:
            print("경고: 진행상황 저장 실패 -", e)

    df = pd.DataFrame(rows)
    # 컬럼 순서 정리
    ordered_cols = [
        "auctionone_court", "auctionone_area", "case_no",
        "region_big", "region_mid", "region_small",
        "usage", "object_group", "address", "address_road", "longitude", "latitude", "coord_crs",
        "auction_date", "big_round",
        "winning_price", "second_big_price",
        "valuation_base_date", "new_build_date", "use_approval_date",
        "building_appraisal_price", "land_appraisal_price",
        "land_area", "building_area",
        "popup_url",
    ]
    # 누락 컬럼 보완
    for c in ordered_cols:
        if c not in df.columns:
            df[c] = None
    return df[ordered_cols], list(processed_pairs_set)


def save_to_excel(df: pd.DataFrame, output_path: str) -> None:
    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)
    # 임시: URL 없는 행 제거(추후 제거 예정)
    before = len(df)
    if "popup_url" in df.columns:
        s = df["popup_url"].astype(str).str.strip()
        mask = df["popup_url"].notna() & s.ne("") & s.str.lower().ne("nan")
        df = df.loc[mask].copy()
    removed = before - len(df)
    if removed > 0:
        print(f"임시 필터 적용: popup_url 비어있는 {removed}건 제거")

    df.to_excel(output_path, index=False)
    # 엑셀에서 링크가 빈 칸처럼 보이는 사례가 있어, 하이퍼링크 속성을 명시 지정
    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        ws = wb.active
        header_idx = {cell.value: cell.column for cell in ws[1]}
        col = header_idx.get("popup_url")
        if col is not None:
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col)
                url = cell.value
                if url and str(url).strip():
                    cell.hyperlink = str(url).strip()
                    cell.style = "Hyperlink"
        wb.save(output_path)
    except Exception as e:
        print("경고: 하이퍼링크 적용 실패 -", e)


# -----------------------------
# 사이트 보강 (Selenium 선로그인 → 상세 페이지 파싱)
# -----------------------------
def _parse_krw(text: str) -> Optional[int]:
    import re as _re
    try:
        if not text:
            return None
        m = _re.search(r"([\d][\d,]*)", str(text))
        if not m:
            return None
        return int(m.group(1).replace(",", ""))
    except Exception:
        return None


def _extract_case_no_from_soup(soup: BeautifulSoup) -> Optional[str]:
    try:
        num = soup.select_one("span.num")
        if not num:
            return None
        core = (num.get_text("", strip=True) or "").strip()
        p = soup.select_one("span.pnum, span.pnum.orange")
        suffix = (p.get_text("", strip=True) or "").strip() if p else ""
        case_no = (core + suffix).strip()
        if "타경" in case_no:
            return case_no
        return None
    except Exception:
        return None


def _extract_address_pair_from_soup(soup: BeautifulSoup) -> Tuple[Optional[str], Optional[str]]:
    """옥션 상세페이지 BeautifulSoup에서 (지번주소, 도로명주소) 한 쌍을 추출한다.

    규칙(data_update와 동일):
      - '소 재 지' tr → 지번(address)
      - '새 주 소' tr → 도로명(address_road)
      - '보관장소' tr → 소재지/도로명 대체용
      - 각 td 안에서 addr_view_1 개수에 따라 주소를 조립
        * 0개: td의 가시 텍스트 전체
        * 1개: td의 가시 텍스트 전체(앞 텍스트 + addr_view_1)
        * 2개 이상: 첫 번째 addr_view_1을 기본주소, 두 번째를 상세로 보고 '기본 + 상세'
    """

    def _norm(s: str) -> str:
        return " ".join((s or "").split())

    def _extract_full_from_td(td) -> Optional[str]:
        if td is None:
            return None
        # 숨김 텍스트/불필요 태그 제거
        for tag in td.select("span.addr_view_0, textarea, script"):
            try:
                tag.decompose()
            except Exception:
                pass
        spans = td.select("span.addr_view_1")
        text_all = _norm(td.get_text(" ", strip=True))
        if not spans:
            return text_all or None
        if len(spans) == 1:
            # 앞 텍스트 + addr_view_1 이 이미 한 줄로 보이므로 전체를 사용
            return text_all or None
        # 2개 이상: 기본 + 상세
        base = _norm(spans[0].get_text(" ", strip=True))
        detail = _norm(spans[1].get_text(" ", strip=True))
        full = f"{base} {detail}".strip()
        return full or text_all or None

    try:
        parcel_td = None     # 소재지
        road_td = None       # 새 주소(도로명)
        storage_td = None    # 보관장소

        for th in soup.find_all("th"):
            try:
                txt = _norm(th.get_text(" ", strip=True))
            except Exception:
                continue
            # '소 재 지'만 인식(공백 없는 '소재지'는 제외)
            if "소 재 지" in txt:
                sib = th.find_next_sibling("td")
                if sib is not None:
                    parcel_td = sib
            # 새주소도 '새 주 소' (띄어쓰기 포함)만 인식
            elif "새 주 소" in txt:
                sib = th.find_next_sibling("td")
                if sib is not None:
                    road_td = sib
            elif ("보관장소" in txt) or ("보관 장소" in txt):
                sib = th.find_next_sibling("td")
                if sib is not None:
                    storage_td = sib

        # 1차: 정석적인 소재지/새주소에서 추출
        addr = _extract_full_from_td(parcel_td)
        addr_road = _extract_full_from_td(road_td)

        # 2차: 소재지가 없고 보관장소만 있는 경우 → 보관장소를 address로 사용
        if addr is None and storage_td is not None:
            addr = _extract_full_from_td(storage_td)

        return (addr, addr_road)
    except Exception:
        return (None, None)


def _extract_prices_from_bidhistory(soup: BeautifulSoup) -> Tuple[Optional[int], Optional[int]]:
    # appraisal_price: '구분=1차' 행의 '최저매각가격'
    # winning_price : '매각 : NNN,NNN,NNN원'
    try:
        import re as _re
        container = soup.select_one("#bidhistory") or soup.select_one("td.tbl_history")
        if not container:
            return (None, None)
        appraisal_price: Optional[int] = None
        winning_price: Optional[int] = None
        table = container.find("table")
        if table:
            rows = table.find_all("tr")
            header_idx = None
            for r in rows:
                ths = r.find_all("th")
                if ths:
                    header_idx = [th.get_text(" ", strip=True) for th in ths]
                    break
            if header_idx:
                try:
                    i_div = header_idx.index("구분")
                except ValueError:
                    i_div = None
                try:
                    i_price = header_idx.index("최저매각가격")
                except ValueError:
                    i_price = None
                data_started = False
                for r in rows:
                    ths = r.find_all("th")
                    if ths and [th.get_text(" ", strip=True) for th in ths] == header_idx:
                        data_started = True
                        continue
                    if not data_started:
                        continue
                    tds = r.find_all(["td", "th"])
                    if not tds:
                        continue
                    if i_div is not None and i_div < len(tds):
                        div_txt = tds[i_div].get_text(" ", strip=True).replace(" ", "")
                        import re as _re
                        if _re.fullmatch(r"(제)?1차", div_txt):
                            if i_price is not None and i_price < len(tds):
                                appraisal_price = _parse_krw(tds[i_price].get_text(" ", strip=True))
                                break
                if appraisal_price is None and data_started:
                    for r in rows:
                        tds = r.find_all("td")
                        if tds and i_price is not None and i_price < len(tds):
                            val = _parse_krw(tds[i_price].get_text(" ", strip=True))
                            if val is not None:
                                appraisal_price = val
                                break
        # 헤더가 없거나 위 로직으로 appraisal_price를 못 찾은 경우:
        #  - 각 행에서 '1차'가 포함된 셀을 찾고, 같은 행에서 '원'이 포함된 금액 텍스트를 금액으로 사용
        if appraisal_price is None and table:
            for r in rows:
                cells = r.find_all(["td", "th"])
                if not cells:
                    continue
                texts = [c.get_text(" ", strip=True).replace(" ", "") for c in cells]
                if not any(_re.fullmatch(r"(제)?1차", t) for t in texts):
                    continue
                # 같은 행에서 첫 번째 금액(숫자+원) 패턴을 찾는다.
                for c in cells:
                    txt = c.get_text(" ", strip=True)
                    m_price = _re.search(r"([\d,]+)\s*원", txt)
                    if m_price:
                        try:
                            appraisal_price = int(m_price.group(1).replace(",", ""))
                        except Exception:
                            appraisal_price = None
                        break
                if appraisal_price is not None:
                    break

        # winning
        cont_text = " ".join(container.stripped_strings)
        m = _re.search(r"매각\s*[:：]\s*([\d,]+)\s*원", cont_text)
        if m:
            try:
                winning_price = int(m.group(1).replace(",", ""))
            except Exception:
                winning_price = None
        else:
            tag = container.find(lambda t: t and t.name in {"span", "article", "section", "div"} and (t.get_text(strip=True) == "매각"))
            if tag and tag.parent:
                ptxt = " ".join(tag.parent.stripped_strings)
                m2 = _re.search(r"매각\s*[:：]\s*([\d,]+)\s*원", ptxt)
                if m2:
                    try:
                        winning_price = int(m2.group(1).replace(",", ""))
                    except Exception:
                        winning_price = None
        return (appraisal_price, winning_price)
    except Exception:
        return (None, None)


def _selenium_start_and_login(login_id: str, login_pw: str, wait_sec: int = 45):
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from webdriver_manager.chrome import ChromeDriverManager
    except Exception as e:
        raise RuntimeError(f"selenium_not_available: {e}")

    service = Service(ChromeDriverManager().install())
    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(service=service, options=options)

    driver.get("https://www.auction1.co.kr/common/login_box.php")
    wait = WebDriverWait(driver, wait_sec)

    try:
        id_field = None
        for css in ["#client_id", "input[name='client_id']", "input#id", "input[name='id']", "input[type='text']"]:
            try:
                el = driver.find_element(By.CSS_SELECTOR, css)
                if el.is_displayed():
                    id_field = el
                    break
            except Exception:
                continue
        pw_field = None
        for css in ["#pw_Dummy", "input[name='pw_Dummy']", "#passwd", "input[name='passwd']", "input[type='password']", "input#password", "input[name='password']"]:
            try:
                el = driver.find_element(By.CSS_SELECTOR, css)
                if el.is_displayed():
                    pw_field = el
                    break
            except Exception:
                continue
        if id_field and pw_field:
            id_field.clear(); id_field.send_keys(login_id)
            if (pw_field.get_attribute("id") or "") == "pw_Dummy":
                pw_field.click(); import time as _t; _t.sleep(0.2)
                real_pw = None
                for sel in ["#passwd", "input[name='passwd']", "input[type='password']", "input#password", "input[name='password']"]:
                    try:
                        el2 = driver.find_element(By.CSS_SELECTOR, sel)
                        if el2.is_displayed():
                            real_pw = el2
                            break
                    except Exception:
                        continue
                (real_pw or pw_field).clear(); (real_pw or pw_field).send_keys(login_pw)
            else:
                pw_field.clear(); pw_field.send_keys(login_pw)
            # 로그인 버튼
            submitted = False
            for css in ["button[type='submit']", "input[type='submit']", "#login_btn", "button.login", "a.login", "span.btn_box_b.btn_darkblue_1", ".btn_box_b.btn_darkblue_1"]:
                try:
                    btn = driver.find_element(By.CSS_SELECTOR, css)
                    if btn.is_displayed():
                        try:
                            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                        except Exception:
                            pass
                        btn.click(); submitted = True; break
                except Exception:
                    continue
            if not submitted:
                from selenium.webdriver.common.keys import Keys
                pw_field.send_keys(Keys.ENTER)
    except Exception:
        pass

    # 로그인 완료 대기
    import time as _t
    end_t = _t.time() + wait_sec
    while _t.time() < end_t:
        cur = driver.current_url or ""
        if "login_box.php" not in cur:
            break
        _t.sleep(0.3)
    return driver


def enrich_rows_with_site(df: pd.DataFrame, login_id: str, login_pw: str, selenium_wait: int = 45, sleep_sec: float = 0.2) -> pd.DataFrame:
    # 필드 보장
    if "address_road" not in df.columns:
        df["address_road"] = None
    if "appraisal_price" not in df.columns:
        df["appraisal_price"] = None
    if "winning_price" not in df.columns:
        df["winning_price"] = None

    driver = _selenium_start_and_login(login_id, login_pw, wait_sec=int(selenium_wait))
    print("브라우저 로그인 완료")

    from tqdm import tqdm as _tqdm
    for i in _tqdm(list(df.index), desc="Enrich from site", unit="row"):
        try:
            url = str(df.at[i, "popup_url"]).strip()
        except Exception:
            url = ""
        if not url or not url.lower().startswith("http"):
            continue
        try:
            driver.get(url)
            html = driver.page_source or ""
            soup = BeautifulSoup(html, "html.parser")

            # 사건번호
            cno = _extract_case_no_from_soup(soup)
            if cno:
                df.at[i, "case_no"] = cno

            # 주소
            addr_main, addr_road = _extract_address_pair_from_soup(soup)
            if addr_main:
                df.at[i, "address"] = addr_main
            if addr_road:
                df.at[i, "address_road"] = addr_road

            # 감정/낙찰가
            a_price, w_price = _extract_prices_from_bidhistory(soup)
            if a_price is not None:
                df.at[i, "appraisal_price"] = a_price
            if w_price is not None:
                df.at[i, "winning_price"] = w_price

            # 지오코딩: road 우선
            chosen = None
            if str(df.at[i, "address_road"] or "").strip():
                chosen = str(df.at[i, "address_road"]).strip()
            elif str(df.at[i, "address"] or "").strip():
                chosen = str(df.at[i, "address"]).strip()
            if chosen:
                x, y, crs = geocode_address_vworld(chosen, prefer_type="ROAD")
                if x is None or y is None:
                    x, y, crs = geocode_address_vworld(chosen, prefer_type="PARCEL")
                if x is not None:
                    df.at[i, "longitude"] = x
                if y is not None:
                    df.at[i, "latitude"] = y
                if crs:
                    df.at[i, "coord_crs"] = crs
        except Exception:
            continue
        finally:
            import time as _t; _t.sleep(max(0.0, float(sleep_sec)))
    try:
        driver.quit()
    except Exception:
        pass
    return df

def main():
    parser = argparse.ArgumentParser(description="data/total/*/*.json → Excel (전체 또는 미처리만)")
    parser.add_argument("--input_root", default="data/total/", help="JSON 루트 디렉토리 (기본: data/total)")
    parser.add_argument("--output", default="results/auction_construct_delta.xlsx", help="출력 엑셀 경로(기본: 델타)")
    parser.add_argument("--use_progress", action="store_true", help="merge_progress.yaml을 읽어 미처리 폴더만 처리")
    parser.add_argument("--progress_path", default="", help="진행상황 YAML 경로(기본: <input_root>/merge_progress.yaml)")
    parser.add_argument("--update_progress", action="store_true", help="처리 후 진행상황 YAML에 완료로 기록")
    # 사이트 보강 옵션
    parser.add_argument("--enrich_site", action="store_true", help="상세 페이지에서 주소/road주소/감정가/낙찰가/좌표 보강")
    parser.add_argument("--login_id", default=os.environ.get("AUCT_ID", ""), help="옥션 사이트 로그인 ID")
    parser.add_argument("--login_pw", default=os.environ.get("AUCT_PW", ""), help="옥션 사이트 로그인 PW")
    parser.add_argument("--selenium_wait", type=int, default=45, help="셀레니움 로그인/로딩 대기(초)")
    parser.add_argument("--sleep", type=float, default=0.2, help="요청 간 대기(초)")
    args = parser.parse_args()

    df, pairs = construct_dataframe(
        args.input_root,
        use_progress=bool(args.use_progress),
        progress_path=(args.progress_path or None),
        update_progress=bool(args.update_progress),
    )
    if args.enrich_site:
        if not args.login_id or not args.login_pw:
            print("경고: --enrich_site 사용 시 --login_id/--login_pw를 설정하세요 (또는 환경변수 AUCT_ID/AUCT_PW)")
        df = enrich_rows_with_site(df, args.login_id or "", args.login_pw or "", selenium_wait=int(args.selenium_wait), sleep_sec=float(args.sleep))
    save_to_excel(df, args.output)
    print(f"생성 완료: {args.output} (행수={len(df)})")

    

if __name__ == "__main__":
    main()
 

